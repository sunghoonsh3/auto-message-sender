import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Read the TXT file with the list of people who received the message
with open('sent_messages.txt', 'r') as file:
    received_list = file.read().splitlines()

# Read the Excel file using pandas
excel_file = 'phone_data.xlsx'
df = pd.read_excel(excel_file)

# Mark recipients in the DataFrame and set 'Marked' for those who received
df['Received'] = df['Email'].isin(received_list)
df.loc[df['Received'], 'Marked'] = 'Marked'

# Load the workbook using openpyxl to modify the existing file
book = load_workbook(excel_file)
sheet = book.active

# Define a light blue fill color
fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Iterate through the rows and fill the entire row until column 'AE' and mark 'Marked' in the 'Marked' column
for row_idx, marked in enumerate(df['Marked'], start=2):  # Assuming header row is present
    if marked == 'Marked':
        for col_idx in range(1, df.columns.get_loc('Marked') + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill

        cell = sheet.cell(row=row_idx, column=df.columns.get_loc('Marked'))
        cell.value = 'Marked'  # Enter 'Marked' in AE column

book.save('marked_recipients.xlsx')
